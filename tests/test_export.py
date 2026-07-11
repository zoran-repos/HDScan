from pathlib import Path

from openpyxl import load_workbook

from file_archive.db.connection import connect
from file_archive.export.excel import export_catalog_to_excel
from file_archive.scanner.engine import scan_directory
from file_archive.search.query import SearchFilters


def _make_tree(root: Path) -> None:
    (root / "books").mkdir()
    (root / "books" / "novel.epub").write_bytes(b"fake epub content")
    (root / "photo.jpg").write_bytes(b"\xff\xd8\xff\xe0fakejpgbytes")
    (root / "notes.txt").write_text("plain text notes")


def test_export_writes_one_row_per_file_with_categories(tmp_path: Path):
    scan_root = tmp_path / "data"
    scan_root.mkdir()
    _make_tree(scan_root)

    db_path = tmp_path / "db" / "archive.db"
    db_path.parent.mkdir(parents=True)
    conn = connect(db_path)
    scan_directory(conn, scan_root)

    output_path = tmp_path / "report.xlsx"
    result_path = export_catalog_to_excel(conn, output_path)
    assert result_path == output_path
    assert output_path.exists()

    wb = load_workbook(output_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert headers == [
        "Disk", "Folder", "Filename", "Extension", "Category",
        "Size (bytes)", "Size", "Created", "Modified", "Hash",
    ]

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 3

    by_filename = {row[2]: row for row in rows}
    assert by_filename["novel.epub"][4] == "Book"
    assert by_filename["photo.jpg"][4] == "Image"
    assert by_filename["notes.txt"][4] == "Document"

    conn.close()


def test_export_respects_filters(tmp_path: Path):
    scan_root = tmp_path / "data"
    scan_root.mkdir()
    _make_tree(scan_root)

    db_path = tmp_path / "db" / "archive.db"
    db_path.parent.mkdir(parents=True)
    conn = connect(db_path)
    scan_directory(conn, scan_root)

    output_path = tmp_path / "images_only.xlsx"
    export_catalog_to_excel(conn, output_path, SearchFilters(extension="jpg", limit=None))

    wb = load_workbook(output_path)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1
    assert rows[0][2] == "photo.jpg"

    conn.close()

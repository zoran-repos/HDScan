"""Excel (.xlsx) report generation for the file catalog.

This is the primary human-facing deliverable of a scan: one row per file,
with folder, filename, type/category, size, dates, and hash.
"""

from __future__ import annotations

import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from file_archive.categorize import categorize
from file_archive.humanize import human_size
from file_archive.search.query import SearchFilters, export_rows

_HEADERS = [
    "Disk",
    "Folder",
    "Filename",
    "Extension",
    "Category",
    "Size (bytes)",
    "Size",
    "Created",
    "Modified",
    "Hash",
]


def _autosize_columns(ws: Worksheet, rows: list[list[object]]) -> None:
    widths = [len(h) for h in _HEADERS]
    for row in rows:
        for i, value in enumerate(row):
            widths[i] = max(widths[i], len(str(value)) if value is not None else 0)
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 10), 80)


def write_excel_report(rows: list[sqlite3.Row], output_path: Path) -> Path:
    """Write one row per cataloged file to an .xlsx workbook at output_path."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Files"
    ws.append(_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    data_rows: list[list[object]] = []
    for row in rows:
        folder = str(Path(row["path"]).parent)
        extension = row["extension"] or ""
        data_rows.append(
            [
                row["disk_label"] or str(row["disk_id"]),
                folder,
                row["filename"],
                extension,
                categorize(extension),
                row["size_bytes"],
                human_size(row["size_bytes"]),
                row["created_date"] or "",
                row["modified_date"] or "",
                row["hash"] or "",
            ]
        )
        ws.append(data_rows[-1])

    _autosize_columns(ws, data_rows)
    wb.save(output_path)
    return output_path


def export_catalog_to_excel(
    conn: sqlite3.Connection, output_path: Path, filters: SearchFilters | None = None
) -> Path:
    rows = export_rows(conn, filters)
    return write_excel_report(rows, output_path)
